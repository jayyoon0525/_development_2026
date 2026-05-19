import random
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
    return value if value != "" else None


def is_email(value):
    if not isinstance(value, str):
        return False
    value = value.strip()
    if not value:
        return False
    return re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$', value) is not None


def find_excel_in_exe_dir():
    exe_dir = Path(sys.argv[0]).resolve().parent
    xlsx_files = sorted(exe_dir.glob("*.xlsx"))
    if len(xlsx_files) == 1:
        return xlsx_files[0]

    preferred_names = [
        "Newsletter_winner_tracking.xlsx",
        "winner_tracking.xlsx",
        "tracking.xlsx",
    ]
    for name in preferred_names:
        candidate = exe_dir / name
        if candidate.exists():
            return candidate

    if len(xlsx_files) > 1:
        raise FileNotFoundError(
            f"여러 개의 엑셀 파일이 있습니다. 실행 파일과 같은 폴더에 사용할 엑셀 파일을 하나만 두거나 파일 이름을 지정하세요.\n" \
            f"찾은 파일: {[p.name for p in xlsx_files]}"
        )

    raise FileNotFoundError(
        f"실행 파일과 동일한 폴더에 Excel(.xlsx) 파일을 찾을 수 없습니다."
    )


def select_vol38_winners(
    excel_path,
    summary_sheet_name="Summary",
    vol_sheet_name="Sheet2",
    source_col=1,
    target_col=2,
    summary_col=5,
    winners_count=3,
    output_path=None,
):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")

    workbook = load_workbook(excel_path)

    if summary_sheet_name not in workbook.sheetnames:
        raise ValueError(f"'{summary_sheet_name}' 시트가 없습니다. 시트 이름을 확인하세요.")
    if vol_sheet_name not in workbook.sheetnames:
        raise ValueError(f"'{vol_sheet_name}' 시트가 없습니다. 시트 이름을 확인하세요.")

    summary_ws = workbook[summary_sheet_name]
    vol_ws = workbook[vol_sheet_name]

    # Vol.38 B1에는 당첨자 헤더를 설정
    vol_ws.cell(row=1, column=target_col, value="당첨자")

    # Summary E열 기존 당첨자 명단 수집
    existing_names = set()
    for row in range(1, summary_ws.max_row + 1):
        value = normalize_value(summary_ws.cell(row=row, column=summary_col).value)
        if value is not None:
            existing_names.add(value)

    # Vol.38 A열 후보자 수집
    candidates = []
    for row in range(2, vol_ws.max_row + 1):
        value = normalize_value(vol_ws.cell(row=row, column=source_col).value)
        if value is None:
            continue
        if not is_email(value):
            continue
        if value not in existing_names:
            candidates.append((row, value))

    if not candidates:
        raise ValueError("Sheet2 A열에서 중복되지 않은 후보자를 찾을 수 없습니다.")

    # 무작위로 선택
    if len(candidates) <= winners_count:
        winners = random.sample(candidates, len(candidates))
        if len(winners) < winners_count:
            print(f"주의: Sheet2 A열에서 중복되지 않은 후보자 수가 {len(winners)}명뿐입니다.")
    else:
        winners = random.sample(candidates, winners_count)

    # Vol.38 B열에 새로운 당첨자 기록
    for row, name in winners:
        vol_ws.cell(row=row, column=target_col, value=name)

    # Summary E열 뒤에 새로운 당첨자 추가
    last_summary_row = 0
    for row in range(1, summary_ws.max_row + 1):
        if normalize_value(summary_ws.cell(row=row, column=summary_col).value) is not None:
            last_summary_row = row

    for index, (_, name) in enumerate(winners, start=1):
        summary_ws.cell(row=last_summary_row + index, column=summary_col, value=name)

    if output_path is None:
        output_path = excel_path
    else:
        output_path = Path(output_path)
        if output_path.exists() and output_path.is_dir():
            output_path = output_path / excel_path.name

    workbook.save(output_path)
    return [name for _, name in winners], output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Sheet2 시트에서 중복되지 않는 새로운 당첨자 3명을 선택하고 Summary 시트에 추가합니다."
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        help="엑셀 파일 경로. 지정하지 않으면 exe와 같은 폴더에서 자동으로 xlsx 파일을 찾습니다."
    )
    parser.add_argument("--summary-sheet", default="Summary", help="Summary 시트 이름")
    parser.add_argument("--vol-sheet", default="Sheet2", help="새 당첨자 후보가 있는 시트 이름")
    parser.add_argument("--output", help="저장할 엑셀 파일 경로 또는 디렉터리. 미지정 시 원본 덮어쓰기")
    parser.add_argument("--count", type=int, default=3, help="선택할 당첨자 수")
    args = parser.parse_args()

    excel_file = args.excel_file
    if excel_file is None:
        excel_file = find_excel_in_exe_dir()
        print(f"자동 탐지된 엑셀 파일: {excel_file}")

    winners, saved_path = select_vol38_winners(
        excel_path=excel_file,
        summary_sheet_name=args.summary_sheet,
        vol_sheet_name=args.vol_sheet,
        winners_count=args.count,
        output_path=args.output,
    )

    print(f"선택된 당첨자: {winners}")
    print(f"저장된 파일: {saved_path}")
