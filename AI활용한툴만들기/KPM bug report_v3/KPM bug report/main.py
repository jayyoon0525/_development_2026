# -*- coding: utf-8 -*-
"""
KPM Issue Report — IVI issue capture and Excel export.
"""

from __future__ import annotations

import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from tkinter import (
    END,
    BooleanVar,
    Canvas,
    Checkbutton,
    Frame,
    Label,
    Menu,
    Radiobutton,
    StringVar,
    Text,
    Tk,
    filedialog,
    messagebox,
    ttk,
)

import pandas as pd

# Internal keys for environment values (all projects; UI shows a subset per project).
ENV_STORAGE_KEYS = frozenset(
    {
        "apk",
        "cluster",
        "vehicle_bench",
        "vin",
        "db",
        "mu_hw",
        "mu_sw",
        "conbox_hw",
        "conbox_sw",
        "gw_hw",
        "gw_sw",
        "hcp3_hw",
        "hcp3_sw",
        "conmod_hw",
        "conmod_sw",
        "hcp5_hw",
        "hcp5_sw",
        "icc_hw",
        "icc_sw",
        "icas2_hw",
        "icas2_sw",
    }
)

# Environment: line1 (Cluster / Vehicle/Bench / VIN); line2 (MU, Conbox, Conmod, GW, ICAS2, …).
ENV_ENTRY_WIDTH = 14
ENV_VIN_ENTRY_WIDTH = ENV_ENTRY_WIDTH * 2
ENV_LINE2_ENTRY_WIDTH = 6
ENV_LINE2_SW_ENTRY_WIDTH = ENV_ENTRY_WIDTH * 2
ENV_FIELD_GROUP_GAP = 0
# Small gap after each field group on line 1 (Cluster / Vehicle/Bench / VIN).
ENV_LINE1_GROUP_GAP = 6
# Same idea on line 2: between all six HW/SW groups (MIB3, HCP3, ICC, ICC MEB).
ENV_LINE2_GROUP_GAP = 6
ENV_LABEL_TO_ENTRY_GAP = 2

# Shared across all report slots (session): reporter, country, project. Environment + date/time are per-report.
SESSION_KEYS = frozenset({"reporter", "country", "project"})

# Optional keys in default-info JSON applied to every report slot (beyond session).
PER_REPORT_DEFAULT_INFO_KEYS = frozenset(
    {
        "rating",
        "function",
        "function_detail",
        "frequency",
        "precondition",
        "action",
        "observed",
        "expected",
        "recovery",
        "error_rate",
        "contact_info",
        "attach",
        "note",
        "remarks",
        "title",
        "full_text",
        "vehicle",
        "found_date",
        "found_time",
        "known_issue",
        "spec_check",
        "retest_verification",
        "kpm_number",
    }
)

# Shipped beside the script or exe; used when the path field is empty or points to a missing file.
# Excel: Field/Value; Vehicle/Bench table (+ optional Cluster, VIN, Project same row); Function + Function detail.
DEFAULT_INFO_FILENAME = "general_info.xlsx"

PROJECT_IDS = ("MIB3", "HCP3", "ICC", "ICC MEB")
DEFAULT_PROJECT = "MIB3"

# Per project: (storage_key, label) in display order.
PROJECT_ENV_FIELDS: dict[str, list[tuple[str, str]]] = {
    "MIB3": [
        ("cluster", "Cluster"),
        ("vehicle_bench", "Vehicle/Bench"),
        ("vin", "VIN"),
        ("mu_hw", "MU HW"),
        ("mu_sw", "MU SW"),
        ("conbox_hw", "Conbox/OCU3 HW"),
        ("conbox_sw", "Conbox/OCU3 SW"),
        ("gw_hw", "GW HW"),
        ("gw_sw", "GW SW"),
    ],
    "HCP3": [
        ("cluster", "Cluster"),
        ("vehicle_bench", "Vehicle/Bench"),
        ("vin", "VIN"),
        ("hcp3_hw", "HCP3 HW"),
        ("hcp3_sw", "HCP3 SW"),
        ("conmod_hw", "Conmod HW"),
        ("conmod_sw", "Conmod SW"),
        ("hcp5_hw", "HCP5 HW"),
        ("hcp5_sw", "HCP5 SW"),
    ],
    "ICC": [
        ("cluster", "Cluster"),
        ("vehicle_bench", "Vehicle/Bench"),
        ("vin", "VIN"),
        ("icc_hw", "ICC HW"),
        ("icc_sw", "ICC SW"),
        ("conmod_hw", "Conmod HW"),
        ("conmod_sw", "Conmod SW"),
        ("gw_hw", "GW HW"),
        ("gw_sw", "GW SW"),
    ],
    "ICC MEB": [
        ("cluster", "Cluster"),
        ("vehicle_bench", "Vehicle/Bench"),
        ("vin", "VIN"),
        ("icc_hw", "ICC HW"),
        ("icc_sw", "ICC SW"),
        ("conmod_hw", "Conmod HW"),
        ("conmod_sw", "Conmod SW"),
        ("icas2_hw", "ICAS2 HW"),
        ("icas2_sw", "ICAS2 SW"),
    ],
}


def _normalize_project(raw: str | None) -> str:
    p = (raw or "").strip() or DEFAULT_PROJECT
    return p if p in PROJECT_ENV_FIELDS else DEFAULT_PROJECT


def _empty_env_values() -> dict[str, str]:
    return {k: "" for k in ENV_STORAGE_KEYS}

# Excel column order (Saved_at = export time; report serial before KPM; Frequency after Rating).
COLUMNS = [
    "Saved_at",
    "Report No.",
    "KPM",
    "Date",
    "Timestamps",
    "Reporter",
    "Country",
    "Vehicle",
    "Function",
    "Title",
    "Description",
    "Rating",
    "Frequency",
]

# openpyxl column width units (approx. character width per column header in COLUMNS order).
ISSUE_REPORT_EXCEL_COL_WIDTHS: dict[str, float] = {
    "Title": 45,
    "Reporter": 20,
    "Date": 10,
    "Timestamps": 12,
    "Country": 8,
    "Report No.": 11,
    "KPM": 36,
    "Rating": 8,
    "Frequency": 12,
    "Function": 30,
    "Vehicle": 8,
    "Description": 20,
}

DEFAULT_COUNTRY = "KR"
# Dropdown lists KR only; combobox stays editable so other ISO codes can be typed.
COUNTRY_COMBO_VALUES = ("KR",)

TITLE_MAX_LEN = 40
TITLE_PREVIEW_COMBO_WIDTH = 48
TITLE_PREVIEW_LINE_CHARS = 52
REPORTER_ENTRY_WIDTH = 28
TITLE_ENTRY_VISIBLE_WIDTH = 43
FORM_WIDE_WIDTH = 70
DESCRIPTION_MULTILINE_WIDTH = 75
# Tk Text height in lines; default precondition bullets need more than a short box.
PRECONDITION_MULTILINE_HEIGHT = 9
# KPM id / free text (Korean, English, digits); no per-keystroke validate (better for IME).
KPM_TEXT_MAX_LEN = 128
FORM_NARROW_WIDTH = 25
CONTACT_INFO_ENTRY_WIDTH = 40

DEFAULT_PRECONDITION = """- System language: KR
- Backend: 
- Online available 
- BT connection:  
- Background: 
- User 
- Destination: 
- Current location:
"""


def _precondition_bullet_key(line: str) -> str | None:
    """Stable key for a `- ...` bullet (prefix before `:` if present, else whole label)."""
    s = (line or "").strip()
    if not s.startswith("-"):
        return None
    rest = s[1:].lstrip()
    if not rest:
        return None
    if ":" in rest:
        return rest.split(":", 1)[0].strip().lower() or None
    return rest.strip().lower() or None


def _default_precondition_slots() -> list[tuple[str, str]]:
    """Ordered (key, default_line) for each bullet in DEFAULT_PRECONDITION."""
    out: list[tuple[str, str]] = []
    for line in DEFAULT_PRECONDITION.splitlines():
        if not line.strip():
            continue
        k = _precondition_bullet_key(line)
        if k:
            out.append((k, line.rstrip()))
    return out


def _precondition_saved_bullet_map(raw: str) -> dict[str, str]:
    """Map bullet key -> full line as saved (last occurrence wins)."""
    m: dict[str, str] = {}
    for line in (raw or "").splitlines():
        k = _precondition_bullet_key(line)
        if k:
            m[k] = line.rstrip()
    return m


def _merge_precondition_with_defaults(raw: str | None) -> str:
    """Fill missing default bullets with template lines; keep saved lines for matching keys."""
    slots = _default_precondition_slots()
    if not slots:
        return (raw or "").strip() or DEFAULT_PRECONDITION
    saved = _precondition_saved_bullet_map(raw or "")
    default_keys = {k for k, _ in slots}
    merged = [saved.get(k, default_line) for k, default_line in slots]
    extras: list[str] = []
    for line in (raw or "").splitlines():
        if not line.strip():
            continue
        k = _precondition_bullet_key(line)
        if k is None:
            extras.append(line.rstrip())
        elif k not in default_keys:
            extras.append(line.rstrip())
    if extras:
        merged.append("")
        merged.extend(extras)
    return "\n".join(merged)


def _precondition_for_form_display(raw: str | None) -> str:
    """Show saved bullets; missing default slots show DEFAULT_PRECONDITION lines for those slots."""
    s = (raw or "").strip()
    if not s:
        return DEFAULT_PRECONDITION
    return _merge_precondition_with_defaults(s)


DEFAULT_ACTION_STEPS = """1. 
2. 
3. 
4. 
5. """

DEFAULT_OBSERVATION_STEPS = """1. 
2. 
3. 
4. 
5. """

DEFAULT_ATTACHMENT = "trace"

# Single Text widget joins Action + Observation; snapshot splits on this marker.
ACTION_OBSERVATION_SEPARATOR = "\n---\n"


def _join_action_observation_for_display(action: str | None, observed: str | None) -> str:
    a = (action or "").rstrip()
    o = (observed or "").rstrip()
    if a and o:
        return f"{a}{ACTION_OBSERVATION_SEPARATOR}{o}"
    return a or o


def _split_action_observation_combined(s: str) -> tuple[str, str]:
    raw = s or ""
    if ACTION_OBSERVATION_SEPARATOR in raw:
        left, right = raw.split(ACTION_OBSERVATION_SEPARATOR, 1)
        return left.strip(), right.strip()
    return raw.strip(), ""


def _safe_date_for_filename(s: str) -> str:
    raw = (s or "").strip() or datetime.now().strftime("%Y-%m-%d")
    return "".join(c if c not in '<>:"/\\|?*' else "_" for c in raw)


def dated_export_basename(date_str: str) -> str:
    return f"KPM_Issue_Report_{_safe_date_for_filename(date_str)}.xlsx"


def _apply_issue_report_sheet_column_widths(path: Path) -> None:
    """Set column widths on the first (issue data) sheet after pandas export."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    for i, name in enumerate(COLUMNS, start=1):
        w = ISSUE_REPORT_EXCEL_COL_WIDTHS.get(name)
        if w is not None:
            ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)


def _add_kpm_issue_charts(path: Path, df: pd.DataFrame) -> None:
    """Append a ``Charts`` sheet with pie charts: counts by Rating and by Function (saved workbook)."""
    if df is None or df.empty:
        return
    if "Rating" not in df.columns or "Function" not in df.columns:
        return

    from openpyxl import load_workbook
    from openpyxl.chart import PieChart, Reference

    rating_s = df["Rating"].fillna("").astype(str).str.strip()
    rating_s = rating_s.replace("", "(empty)")
    r_counts = rating_s.value_counts()
    rating_order = ["A1", "A", "B", "C"]
    r_items: list[tuple[str, int]] = []
    seen_r: set[str] = set()
    for lab in rating_order:
        if lab in r_counts.index:
            r_items.append((lab, int(r_counts.loc[lab])))
            seen_r.add(lab)
    for lab in sorted(r_counts.index, key=lambda x: str(x)):
        if lab not in seen_r:
            r_items.append((str(lab), int(r_counts.loc[lab])))

    fn_s = df["Function"].fillna("").astype(str).str.strip()
    fn_s = fn_s.replace("", "(empty)")
    f_counts = fn_s.value_counts().sort_values(ascending=False)
    f_items = [(str(lab)[:250], int(f_counts.loc[lab])) for lab in f_counts.index]

    wb = load_workbook(path)
    chart_name = "Charts"
    if chart_name in wb.sheetnames:
        wb.remove(wb[chart_name])
    ws = wb.create_sheet(chart_name)

    ws["A1"], ws["B1"] = "Rating", "Count"
    row = 2
    for lab, cnt in r_items:
        ws.cell(row=row, column=1, value=lab)
        ws.cell(row=row, column=2, value=cnt)
        row += 1
    r_last = row - 1

    if r_last >= 2:
        ch_r = PieChart()
        ch_r.title = "Issues by Rating"
        data = Reference(ws, min_col=2, min_row=1, max_row=r_last)
        cats = Reference(ws, min_col=1, min_row=2, max_row=r_last)
        ch_r.add_data(data, titles_from_data=True)
        ch_r.set_categories(cats)
        ws.add_chart(ch_r, "D2")

    fn_start = max(r_last + 10, 20)
    ws.cell(row=fn_start, column=1, value="Function")
    ws.cell(row=fn_start, column=2, value="Count")
    row = fn_start + 1
    for lab, cnt in f_items:
        ws.cell(row=row, column=1, value=lab)
        ws.cell(row=row, column=2, value=cnt)
        row += 1
    f_last = row - 1

    if f_last >= fn_start + 1:
        ch_f = PieChart()
        ch_f.title = "Issues by Function"
        data = Reference(ws, min_col=2, min_row=fn_start, max_row=f_last)
        cats = Reference(ws, min_col=1, min_row=fn_start + 1, max_row=f_last)
        ch_f.add_data(data, titles_from_data=True)
        ch_f.set_categories(cats)
        ws.add_chart(ch_f, f"D{fn_start}")

    ws.column_dimensions["A"].width = 48
    wb.save(path)


def _excel_function_function_detail(d: dict[str, str]) -> str:
    """Excel Function column: ``Function-Function detail`` when both set; otherwise whichever is set."""
    fn = (d.get("function") or "").strip()
    fd = (d.get("function_detail") or "").strip()
    if fn and fd:
        return f"{fn}-{fd}"
    return fn or fd


def _combined_action_and_observation(d: dict[str, str]) -> str:
    a = (d.get("action") or "").strip()
    o = (d.get("observed") or "").strip()
    if a and o:
        return f"{a}\n{o}"
    return a or o


def _flag_truthy(raw: str | None) -> bool:
    return (raw or "").strip().lower() in ("1", "true", "yes", "y", "on")


def _format_kpm_excel_cell(d: dict[str, str]) -> str:
    """Excel ``KPM`` column: KPM id plus checked issue flags (labels as requested)."""
    parts: list[str] = []
    kpm = (d.get("kpm_number") or "").strip()
    if kpm:
        parts.append(kpm)
    if _flag_truthy(d.get("known_issue")):
        parts.append("Known issue()")
    if _flag_truthy(d.get("spec_check")):
        parts.append("spec check")
    if _flag_truthy(d.get("retest_verification")):
        parts.append("retest/verification")
    return "; ".join(parts)


def resource_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def bundled_default_info_path() -> Path:
    return resource_path() / DEFAULT_INFO_FILENAME


def resolve_general_info_excel_path(raw: str | None) -> Path:
    """Use the path from the UI if it exists; otherwise fall back to general_info beside the app."""
    bundled = bundled_default_info_path()
    s = (raw or "").strip()
    if not s:
        return bundled
    try:
        p = Path(s).expanduser()
        if p.is_file():
            return p.resolve()
    except OSError:
        pass
    return bundled


def ensure_bundled_general_info_excel() -> None:
    """Create an empty Field/Value template next to the app if the bundled file is missing."""
    path = bundled_default_info_path()
    if path.is_file():
        return
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame({"Field": [], "Value": [], "VIN": []}).to_excel(
            path, index=False, engine="openpyxl"
        )
    except Exception:
        pass


def _excel_cell_to_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float):
        if pd.isna(v):
            return ""
        if v == int(v):
            return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _default_info_field_normalize(field: str) -> str:
    """Lowercase alphanumerics only (so Vehicle/Bench, vehicle_bench, Vehicle-Bench match)."""
    return "".join(c.lower() for c in (field or "").strip() if c.isalnum())


def _strip_excel_text(s: object) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    t = str(s).replace("\ufeff", "").strip()
    return t


def _excel_column_key_norm(col: object) -> str:
    return _default_info_field_normalize(_strip_excel_text(col))


DEFAULT_INFO_ALLOWED_KEYS = SESSION_KEYS | PER_REPORT_DEFAULT_INFO_KEYS | ENV_STORAGE_KEYS


def _canonical_default_info_field_key(raw_field: object) -> str | None:
    """Map Excel Field column text to internal storage keys.

    Headers like ``Function`` or ``Cluster`` must match despite casing; the allowed-key
    set uses lowercase identifiers only.
    """
    key = _strip_excel_text(raw_field)
    if not key:
        return None
    nk = _default_info_field_normalize(key)
    if nk == "vehiclebenchoptions":
        return None
    if nk == "vehiclebench":
        return "vehicle_bench"
    if nk == "vin":
        return "vin"
    for allowed in DEFAULT_INFO_ALLOWED_KEYS:
        if _default_info_field_normalize(allowed) == nk:
            return allowed
    return None


def _vehicle_table_column_map(df: pd.DataFrame) -> dict[str, int] | None:
    """Map bench / cluster / vin / project to column indices. None if not a vehicle list table."""
    if df.shape[1] < 2 or df.empty:
        return None
    m: dict[str, int] = {}
    for i, col in enumerate(df.columns):
        n = _excel_column_key_norm(col)
        if n in ("vehiclebench", "bench"):
            m["bench"] = i
        elif n == "vehicle":
            m.setdefault("bench", i)
        elif n == "cluster":
            m["cluster"] = i
        elif n == "vin":
            m["vin"] = i
        elif n == "project":
            m["project"] = i
    if "bench" not in m:
        return None
    c0 = _excel_column_key_norm(df.columns[0])
    c1 = _excel_column_key_norm(df.columns[1]) if len(df.columns) > 1 else ""
    if len(df.columns) >= 2:
        if c0 == "vehiclebench" and c1 == "vin":
            m["bench"], m["vin"] = 0, 1
        elif c0 == "vehiclebench" and c1 == "cluster":
            m["bench"], m["cluster"] = 0, 1
        elif c0 == "bench" and c1 == "vin":
            m["bench"], m["vin"] = 0, 1
    if (
        "vin" not in m
        and "cluster" not in m
        and "project" not in m
        and len(df.columns) == 2
        and m.get("bench") == 0
    ):
        m["vin"] = 1
    if "vin" not in m and "cluster" not in m and "project" not in m:
        return None
    return m


def _is_vehicle_bench_data_table(df: pd.DataFrame) -> bool:
    return _vehicle_table_column_map(df) is not None


# Set when pd.read_excel fails (e.g. file locked, missing DLL in frozen build); cleared on success.
_last_general_info_read_error: str | None = None


def _read_excel_all_sheets(path: Path) -> dict[str, pd.DataFrame]:
    global _last_general_info_read_error
    if not path.is_file():
        _last_general_info_read_error = f"File not found: {path}"
        return {}
    try:
        raw = pd.read_excel(path, engine="openpyxl", sheet_name=None)
        _last_general_info_read_error = None
    except Exception as e:
        _last_general_info_read_error = f"{type(e).__name__}: {e}"
        return {}
    if isinstance(raw, dict):
        return raw
    return {"Sheet1": raw}


def _is_vehicle_bench_single_column_table(df: pd.DataFrame) -> bool:
    """One column of bench names (header: Vehicle/Bench, Bench, or Vehicle)."""
    if df is None or df.empty or df.shape[1] != 1:
        return False
    c0 = _excel_column_key_norm(df.columns[0])
    return c0 in ("vehiclebench", "bench", "vehicle")


def _read_vehicle_bench_single_column_df(df: pd.DataFrame) -> tuple[list[str], dict[str, dict[str, str]]]:
    if not _is_vehicle_bench_single_column_table(df):
        return [], {}
    out: list[str] = []
    attrs: dict[str, dict[str, str]] = {}
    seen: set[str] = set()
    for _, row in df.iterrows():
        cell = row.iloc[0]
        if pd.isna(cell):
            continue
        b = _excel_cell_to_str(cell).strip()
        if not b:
            continue
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out, attrs


def _is_function_detail_table(df: pd.DataFrame) -> bool:
    if df.shape[1] < 2 or df.empty:
        return False
    c0 = _excel_column_key_norm(df.columns[0])
    c1 = _excel_column_key_norm(df.columns[1])
    if c0 != "function":
        return False
    # "Function" + "Function detail" → functiondetail; some sheets use only "Detail".
    return c1 in ("functiondetail", "detail")


def _read_function_detail_from_table(df: pd.DataFrame) -> tuple[list[str], dict[str, list[str]]]:
    fn_order: list[str] = []
    seen_fn: set[str] = set()
    details: dict[str, list[str]] = {}
    for _, row in df.iterrows():
        f_cell = row.iloc[0]
        d_cell = row.iloc[1] if df.shape[1] >= 2 else None
        if pd.isna(f_cell):
            continue
        f = _excel_cell_to_str(f_cell).strip()
        if not f:
            continue
        d_str = ""
        if d_cell is not None and not (isinstance(d_cell, float) and pd.isna(d_cell)):
            d_str = _excel_cell_to_str(d_cell).strip()
        if f not in seen_fn:
            seen_fn.add(f)
            fn_order.append(f)
        if f not in details:
            details[f] = []
        if d_str and d_str not in details[f]:
            details[f].append(d_str)
    return fn_order, details


def _read_function_sync_from_excel(path: Path) -> tuple[list[str], dict[str, list[str]]]:
    """Collect Function / Function detail rows from every sheet that matches the table shape."""
    order: list[str] = []
    details: dict[str, list[str]] = {}
    seen_fn: set[str] = set()
    for df in _read_excel_all_sheets(path).values():
        if df is None or df.empty or len(df.columns) < 2:
            continue
        if not _is_function_detail_table(df):
            continue
        o, d = _read_function_detail_from_table(df)
        for f in o:
            if f not in seen_fn:
                seen_fn.add(f)
                order.append(f)
        for f, lst in d.items():
            if f not in details:
                details[f] = []
            for x in lst:
                if x not in details[f]:
                    details[f].append(x)
    return order, details


def _read_vehicle_bench_from_table(df: pd.DataFrame) -> tuple[list[str], dict[str, dict[str, str]]]:
    cmap = _vehicle_table_column_map(df)
    if cmap is None:
        return [], {}
    out: list[str] = []
    attrs: dict[str, dict[str, str]] = {}
    seen: set[str] = set()
    bi = cmap["bench"]
    for _, row in df.iterrows():
        bench_cell = row.iloc[bi]
        if pd.isna(bench_cell):
            continue
        b = _excel_cell_to_str(bench_cell).strip()
        if not b:
            continue
        if b not in seen:
            seen.add(b)
            out.append(b)
        row_a: dict[str, str] = {}
        if "cluster" in cmap:
            cc = row.iloc[cmap["cluster"]]
            if not (isinstance(cc, float) and pd.isna(cc)):
                cs = _excel_cell_to_str(cc).strip()
                if cs:
                    row_a["cluster"] = cs
        if "vin" in cmap:
            vc = row.iloc[cmap["vin"]]
            if not (isinstance(vc, float) and pd.isna(vc)):
                vs = _excel_cell_to_str(vc).strip()
                if vs:
                    row_a["vin"] = vs
        if "project" in cmap:
            pc = row.iloc[cmap["project"]]
            if not (isinstance(pc, float) and pd.isna(pc)):
                ps = _excel_cell_to_str(pc).strip()
                if ps:
                    row_a["project"] = _normalize_project(ps)
        attrs[b] = row_a
    return out, attrs


def _read_default_info_from_dataframe(df: pd.DataFrame) -> dict[str, str]:
    if df.empty or len(df.columns) < 2:
        return {}
    if _is_vehicle_bench_data_table(df) or _is_function_detail_table(df):
        return {}
    kcol, vcol = df.columns[0], df.columns[1]
    out: dict[str, str] = {}
    for _, row in df.iterrows():
        raw_k = row[kcol]
        if pd.isna(raw_k):
            continue
        key = _canonical_default_info_field_key(raw_k)
        if key is None:
            continue
        raw_v = row[vcol]
        out[key] = "" if pd.isna(raw_v) else _excel_cell_to_str(raw_v)
    return out


def _read_default_info_from_excel(path: Path) -> dict[str, str]:
    """Load Field/Value defaults from every sheet that is not a Vehicle/Bench or Function table."""
    merged: dict[str, str] = {}
    for df in _read_excel_all_sheets(path).values():
        if df is None or df.empty or len(df.columns) < 2:
            continue
        part = _read_default_info_from_dataframe(df)
        merged.update(part)
    return merged


def _read_vehicle_bench_field_value_df(df: pd.DataFrame) -> tuple[list[str], dict[str, dict[str, str]]]:
    if df.empty or len(df.columns) < 2:
        return [], {}
    if _is_vehicle_bench_data_table(df) or _is_function_detail_table(df):
        return [], {}
    kcol, vcol = df.columns[0], df.columns[1]
    vin_col = df.columns[2] if len(df.columns) >= 3 else None
    cluster_col = df.columns[3] if len(df.columns) >= 4 else None
    project_col = df.columns[4] if len(df.columns) >= 5 else None
    out: list[str] = []
    attrs: dict[str, dict[str, str]] = {}
    seen: set[str] = set()

    def add_one(s: str) -> None:
        t = (s or "").strip()
        if t and t not in seen:
            seen.add(t)
            out.append(t)

    def set_attr(bench: str, key: str, cell: object) -> None:
        b = (bench or "").strip()
        if not b or cell is None or (isinstance(cell, float) and pd.isna(cell)):
            return
        vs = _excel_cell_to_str(cell).strip()
        if not vs:
            return
        if key == "project":
            vs = _normalize_project(vs)
        if b not in attrs:
            attrs[b] = {}
        attrs[b][key] = vs

    for _, row in df.iterrows():
        raw_k = row[kcol]
        if pd.isna(raw_k):
            continue
        nk = _default_info_field_normalize(_strip_excel_text(raw_k))
        raw_v = row[vcol]
        if nk == "vehiclebenchoptions":
            if pd.isna(raw_v):
                continue
            benches = [x.strip() for x in str(raw_v).split(",") if x.strip()]
            vins: list[str] = []
            clusters: list[str] = []
            projects: list[str] = []
            if vin_col is not None:
                raw_vin = row[vin_col]
                if not pd.isna(raw_vin):
                    vins = [x.strip() for x in str(raw_vin).split(",")]
            if cluster_col is not None:
                raw_c = row[cluster_col]
                if not pd.isna(raw_c):
                    clusters = [x.strip() for x in str(raw_c).split(",")]
            if project_col is not None:
                raw_p = row[project_col]
                if not pd.isna(raw_p):
                    projects = [x.strip() for x in str(raw_p).split(",")]
            for i, b in enumerate(benches):
                add_one(b)
                if i < len(vins) and vins[i]:
                    set_attr(b, "vin", vins[i])
                if i < len(clusters) and clusters[i]:
                    set_attr(b, "cluster", clusters[i])
                if i < len(projects) and projects[i]:
                    set_attr(b, "project", projects[i])
            continue
        if nk != "vehiclebench":
            continue
        if pd.isna(raw_v):
            continue
        val = _excel_cell_to_str(raw_v)
        if val:
            add_one(val)
            if vin_col is not None:
                set_attr(val, "vin", row[vin_col])
            if cluster_col is not None:
                set_attr(val, "cluster", row[cluster_col])
            if project_col is not None:
                set_attr(val, "project", row[project_col])
    return out, attrs


def _read_vehicle_bench_sync_from_excel(path: Path) -> tuple[list[str], dict[str, dict[str, str]]]:
    """Merge Vehicle/Bench rows with optional cluster + VIN per bench from all matching sheets."""
    sheets = _read_excel_all_sheets(path)
    if not sheets:
        return [], {}
    out: list[str] = []
    bench_attrs: dict[str, dict[str, str]] = {}
    seen: set[str] = set()

    def append_benches(order: list[str], mapping: dict[str, dict[str, str]]) -> None:
        for b in order:
            if b not in seen:
                seen.add(b)
                out.append(b)
        for b, a in mapping.items():
            if b not in bench_attrs:
                bench_attrs[b] = {}
            bench_attrs[b].update(a)

    for df in sheets.values():
        if df is None or df.empty:
            continue
        if df.shape[1] == 1 and _is_vehicle_bench_single_column_table(df):
            o, m = _read_vehicle_bench_single_column_df(df)
            append_benches(o, m)

    for df in sheets.values():
        if df is None or df.empty or len(df.columns) < 2:
            continue
        if _is_vehicle_bench_data_table(df):
            o, m = _read_vehicle_bench_from_table(df)
            append_benches(o, m)

    for df in sheets.values():
        if df is None or df.empty or len(df.columns) < 2:
            continue
        if _is_vehicle_bench_data_table(df) or _is_function_detail_table(df):
            continue
        o, m = _read_vehicle_bench_field_value_df(df)
        if o or m:
            append_benches(o, m)

    return out, bench_attrs


def _env_val(d: dict[str, str], key: str) -> str:
    v = (d.get(key) or "").strip()
    return v if v else "(none)"


def _env_line_from_pairs(d: dict[str, str], pairs: list[tuple[str, str]]) -> str:
    """One Excel line: 'Label: v1, Label: v2' for HW/SW groups."""
    return ", ".join(f"{label}: {_env_val(d, key)}" for key, label in pairs)


def _format_environment_excel_lines(proj: str, d: dict[str, str]) -> list[str]:
    """Environment block for Excel: project/cluster and HW+SW pairs comma-separated per line."""
    lines: list[str] = [
        f"Project: {proj}, Cluster: {_env_val(d, 'cluster')}",
        _env_line_from_pairs(d, [("vehicle_bench", "Vehicle/Bench")]),
        f"VIN: {_env_val(d, 'vin')}",
        f"DB: {_env_val(d, 'db')}, APK: {_env_val(d, 'apk')}",
    ]
    if proj == "MIB3":
        lines.extend(
            [
                _env_line_from_pairs(d, [("mu_hw", "MU HW"), ("mu_sw", "MU SW")]),
                _env_line_from_pairs(
                    d, [("conbox_hw", "Conbox/OCU3 HW"), ("conbox_sw", "Conbox/OCU3 SW")]
                ),
                _env_line_from_pairs(d, [("gw_hw", "GW HW"), ("gw_sw", "GW SW")]),
            ]
        )
    elif proj == "HCP3":
        lines.extend(
            [
                _env_line_from_pairs(d, [("hcp3_hw", "HCP3 HW"), ("hcp3_sw", "HCP3 SW")]),
                _env_line_from_pairs(
                    d, [("conmod_hw", "Conmod HW"), ("conmod_sw", "Conmod SW")]
                ),
                _env_line_from_pairs(d, [("hcp5_hw", "HCP5 HW"), ("hcp5_sw", "HCP5 SW")]),
            ]
        )
    elif proj == "ICC":
        lines.extend(
            [
                _env_line_from_pairs(d, [("icc_hw", "ICC HW"), ("icc_sw", "ICC SW")]),
                _env_line_from_pairs(
                    d, [("conmod_hw", "Conmod HW"), ("conmod_sw", "Conmod SW")]
                ),
                _env_line_from_pairs(d, [("gw_hw", "GW HW"), ("gw_sw", "GW SW")]),
            ]
        )
    elif proj == "ICC MEB":
        lines.extend(
            [
                _env_line_from_pairs(d, [("icc_hw", "ICC HW"), ("icc_sw", "ICC SW")]),
                _env_line_from_pairs(
                    d, [("conmod_hw", "Conmod HW"), ("conmod_sw", "Conmod SW")]
                ),
                _env_line_from_pairs(
                    d, [("icas2_hw", "ICAS2 HW"), ("icas2_sw", "ICAS2 SW")]
                ),
            ]
        )
    lines.append("")
    lines.append(f"Frequency: {_env_val(d, 'frequency')}")
    return lines


def _format_description_excel(d: dict[str, str]) -> str:
    """Single Description column: tagged sections; [Full Text] first; no issue-flag block (KPM is its own Excel column)."""

    def blk(tag: str, body: str) -> str:
        t = (body or "").strip()
        return f"{tag}\n{t if t else '(none)'}"

    proj = _normalize_project(d.get("project"))
    env_body = "\n".join(_format_environment_excel_lines(proj, d))

    parts = [
        blk("[Full Text]", d.get("full_text", "")),
        blk("[Date]", d.get("found_date", "")),
        blk("[Timestamp]", d.get("found_time", "")),
        f"[Environment]\n{env_body}",
        blk("[Precondition]", d.get("precondition", "")),
        blk("[Action and Observation]", _combined_action_and_observation(d)),
        blk("[Note]", d.get("note", "") or d.get("remarks", "")),
        blk("[Expected behavior]", d.get("expected", "")),
        blk("[Recovery]", d.get("recovery", "")),
        blk("[Error rate]", d.get("error_rate", "")),
        blk("[Attachment]", d.get("attach", "")),
        blk("[Reporter]", d.get("reporter", "")),
        blk("[Contact info]", d.get("contact_info", "")),
    ]
    return "\n\n".join(parts)


class BugReportApp:
    def __init__(self) -> None:
        self.root = Tk()
        self.root.title("KPM Issue Report")
        self.root.minsize(720, 560)
        self.root.geometry("1030x640")

        self.excel_path = StringVar()
        self.append_mode = BooleanVar(value=False)
        self.default_info_path = StringVar(value=str(bundled_default_info_path()))

        self._reports: list[dict[str, str]] = []
        self._current_idx = 0
        self._scroll_canvas: Canvas | None = None
        self._scroll_form: Frame | None = None
        self._session_common: dict[str, str] = {}
        # Debounce canvas scroll-on-focus (FocusIn storms + update_idletasks hurt IME typing).
        self._focus_scroll_after_id: str | None = None

        self._build_menu()
        self._build_nav_bar()
        self._build_scrollable_form()
        self._build_actions()

        self._reports = [self._default_report_dict()]
        self._current_idx = 0
        self._session_common = {k: self._reports[0].get(k, "") for k in SESSION_KEYS}
        if not (self._session_common.get("country") or "").strip():
            self._session_common["country"] = DEFAULT_COUNTRY
        if not (self._reports[0].get("vehicle_bench") or "").strip():
            leg = (self._reports[0].get("vehicle") or "").strip()
            if leg:
                self._reports[0]["vehicle_bench"] = leg
        self._merge_default_info_from_file(show_read_error=True)
        self._apply_to_form(self._reports[0])
        self._sync_vehicle_row_from_general_info()
        self._set_excel_path_for_current_date()
        self.var_found_date.trace_add("write", lambda *_: self._set_excel_path_for_current_date())
        self._update_nav_label()

    def _default_report_dict(self) -> dict[str, str]:
        now = datetime.now()
        d: dict[str, str] = {
            "title": "",
            "full_text": "",
            "reporter": "",
            "found_date": now.strftime("%Y-%m-%d"),
            "found_time": now.strftime("%H:%M"),
            "rating": "A",
            "function": "",
            "function_detail": "",
            "project": DEFAULT_PROJECT,
            "precondition": DEFAULT_PRECONDITION,
            "frequency": "Always",
            "action": DEFAULT_ACTION_STEPS,
            "observed": DEFAULT_OBSERVATION_STEPS,
            "expected": "",
            "recovery": "",
            "error_rate": "",
            "contact_info": "",
            "attach": DEFAULT_ATTACHMENT,
            "note": "",
            "known_issue": "",
            "spec_check": "",
            "retest_verification": "",
            "kpm_number": "",
        }
        d.update(_empty_env_values())
        return d

    def _blank_per_report_dict(self) -> dict[str, str]:
        """Fields that are independent per report slot (no session keys)."""
        return {
            "title": "",
            "full_text": "",
            "rating": "A",
            "function": "",
            "function_detail": "",
            "precondition": DEFAULT_PRECONDITION,
            "frequency": "Always",
            "action": DEFAULT_ACTION_STEPS,
            "observed": DEFAULT_OBSERVATION_STEPS,
            "expected": "",
            "recovery": "",
            "error_rate": "",
            "contact_info": "",
            "attach": DEFAULT_ATTACHMENT,
            "note": "",
            "known_issue": "",
            "spec_check": "",
            "retest_verification": "",
            "kpm_number": "",
        }

    def _merge_session(self, d: dict[str, str]) -> dict[str, str]:
        out = dict(d)
        out.update(self._session_common)
        if not (out.get("vehicle_bench") or "").strip() and (out.get("vehicle") or "").strip():
            out = dict(out)
            out["vehicle_bench"] = (out.get("vehicle") or "").strip()
        return out

    def _sync_common_from_form(self) -> None:
        proj = _normalize_project(self.var_project.get())
        self._session_common = {
            "reporter": self.var_reporter.get().strip(),
            "country": self.var_country.get().strip() or DEFAULT_COUNTRY,
            "project": proj,
        }

    def _apply_common_to_form(self) -> None:
        self.var_reporter.set(self._session_common.get("reporter", ""))
        self.var_country.set(self._session_common.get("country", "") or DEFAULT_COUNTRY)
        self.var_project.set(_normalize_project(self._session_common.get("project")))

    def _new_report_dict(self) -> dict[str, str]:
        self._sync_common_from_form()
        d = self._blank_per_report_dict()
        d.update(_empty_env_values())
        d.update(self._session_common)
        now = datetime.now()
        d["found_date"] = now.strftime("%Y-%m-%d")
        d["found_time"] = now.strftime("%H:%M")
        return d

    def _export_parent_dir(self) -> Path:
        raw = (self.excel_path.get() or "").strip()
        if not raw:
            return resource_path()
        p = Path(raw)
        if p.suffix.lower() == ".xlsx":
            par = p.parent
        else:
            par = p if p.is_dir() else p.parent
        try:
            return par.resolve()
        except Exception:
            return resource_path()

    def _set_excel_path_for_current_date(self) -> None:
        if not hasattr(self, "var_found_date"):
            return
        d = self.var_found_date.get().strip() or datetime.now().strftime("%Y-%m-%d")
        parent = self._export_parent_dir()
        try:
            parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            parent = resource_path()
        self.excel_path.set(str(parent / dated_export_basename(d)))

    def _unique_path(self, path: Path) -> Path:
        if not path.exists():
            return path
        stem, suf = path.stem, path.suffix
        n = 2
        while True:
            cand = path.parent / f"{stem}_{n}{suf}"
            if not cand.exists():
                return cand
            n += 1

    def _resolve_save_path(self) -> Path:
        self._sync_common_from_form()
        date_str = (self.var_found_date.get() or "").strip() or datetime.now().strftime("%Y-%m-%d")
        raw = (self.excel_path.get() or "").strip()
        p = Path(raw) if raw else None

        if self.append_mode.get():
            if p and p.suffix.lower() == ".xlsx":
                p.parent.mkdir(parents=True, exist_ok=True)
                return p
            parent = self._export_parent_dir()
            parent.mkdir(parents=True, exist_ok=True)
            return parent / dated_export_basename(date_str)

        parent = self._export_parent_dir()
        parent.mkdir(parents=True, exist_ok=True)
        base = parent / dated_export_basename(date_str)
        return self._unique_path(base)

    def _build_menu(self) -> None:
        menubar = Menu(self.root)
        file_menu = Menu(menubar, tearoff=0)
        file_menu.add_command(label="Choose Excel file…", command=self._pick_excel)
        file_menu.add_command(
            label="Reload general_info (sync defaults)…",
            command=self._reload_general_info_from_excel,
        )
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self._quit_app)
        menubar.add_cascade(label="File", menu=file_menu)
        self.root.config(menu=menubar)

    def _build_nav_bar(self) -> None:
        bar = Frame(self.root, padx=8, pady=4)
        bar.pack(fill="x")
        try:
            _sty = ttk.Style(self.root)
            _sty.configure("NavPrevMuted.TButton", foreground="#6d6d6d")
        except Exception:
            pass
        self._btn_nav_prev = ttk.Button(
            bar, text="« Previous", command=self._on_nav_prev_clicked, style="TButton"
        )
        self._btn_nav_prev.pack(side="left")
        self._nav_label_var = StringVar(value="Report 1 / 1")
        Label(bar, textvariable=self._nav_label_var, padx=16).pack(side="left")
        btn_next = ttk.Button(bar, text="Next »", command=self._nav_next)
        btn_next.pack(side="left")
        self._bind_enter_runs_command(btn_next, self._nav_next)
        self._cb_append_rows = Checkbutton(
            bar,
            text="Append rows to existing file",
            variable=self.append_mode,
            takefocus=1,
        )
        self._cb_append_rows.pack(side="left", padx=(12, 0))
        self._cb_append_rows.bind("<Return>", self._on_append_checkbox_return)
        self._cb_append_rows.bind("<KP_Enter>", self._on_append_checkbox_return)

    def _build_excel_bar(self, parent: Frame) -> None:
        bar = Frame(parent)
        bar.pack(fill="x", pady=(0, 8))
        path_row = Frame(bar)
        path_row.pack(fill="x")
        Label(path_row, text="Save path:").pack(side="left")
        entry = ttk.Entry(path_row, textvariable=self.excel_path, width=80)
        entry.pack(side="left", padx=(6, 6))
        btn_browse_excel = ttk.Button(path_row, text="Browse…", command=self._pick_excel)
        btn_browse_excel.pack(side="left")
        self._bind_enter_runs_command(btn_browse_excel, self._pick_excel)

        info_row = Frame(bar)
        info_row.pack(fill="x", pady=(4, 0))
        Label(info_row, text="Default info path:").pack(side="left")
        ttk.Entry(info_row, textvariable=self.default_info_path, width=80).pack(
            side="left", padx=(6, 6)
        )
        btn_browse_default = ttk.Button(
            info_row, text="Browse…", command=self._pick_default_info_path
        )
        btn_browse_default.pack(side="left")
        self._bind_enter_runs_command(btn_browse_default, self._pick_default_info_path)

    def _refresh_scrollregion(self) -> None:
        c = self._scroll_canvas
        if c is None:
            return
        c.update_idletasks()
        bbox = c.bbox("all")
        if bbox:
            c.configure(scrollregion=bbox)

    def _unpost_all_ttk_comboboxes(self, parent=None, depth: int = 0) -> None:
        """Close any open ttk Combobox lists before scrolling the canvas (avoids detached popups on Windows)."""
        if depth > 50:
            return
        if parent is None:
            parent = self._scroll_form
        if parent is None:
            return
        for child in parent.winfo_children():
            try:
                if child.winfo_class() == "TCombobox":
                    self.root.tk.call("ttk::combobox::Unpost", child)
            except Exception:
                pass
            self._unpost_all_ttk_comboboxes(child, depth + 1)

    def _do_canvas_wheel_delta(self, delta: int) -> None:
        c = self._scroll_canvas
        if c is None or delta == 0:
            return
        self._unpost_all_ttk_comboboxes()
        steps = int(-delta / 120)
        if steps == 0:
            steps = -1 if delta > 0 else 1
        c.yview_scroll(steps, "units")
        self.root.after_idle(self._refresh_scrollregion)

    def _do_canvas_wheel_linux(self, direction: int) -> None:
        c = self._scroll_canvas
        if c is None:
            return
        self._unpost_all_ttk_comboboxes()
        c.yview_scroll(direction, "units")
        self.root.after_idle(self._refresh_scrollregion)

    def _pointer_inside_canvas(self, xr: int, yr: int) -> bool:
        c = self._scroll_canvas
        if c is None:
            return False
        try:
            cx, cy = c.winfo_rootx(), c.winfo_rooty()
            cw, ch = c.winfo_width(), c.winfo_height()
        except Exception:
            return False
        if cw <= 1 or ch <= 1:
            return False
        return cx <= xr < cx + cw and cy <= yr < cy + ch

    def _suppress_canvas_wheel(self, event) -> bool:
        """Do not scroll the main form while a separate toplevel has focus (menus, dialogs)."""
        w = getattr(event, "widget", None)
        if w is not None:
            try:
                return w.winfo_toplevel() is not self.root
            except Exception:
                pass
        return False

    def _wheel_scroll(self, event) -> None:
        xr, yr = event.x_root, event.y_root
        if not self._pointer_inside_canvas(xr, yr):
            return
        if self._suppress_canvas_wheel(event):
            return
        delta = getattr(event, "delta", 0) or 0
        if delta == 0:
            return
        self._do_canvas_wheel_delta(delta)

    def _wheel_scroll_linux(self, event, direction: int) -> None:
        xr, yr = event.x_root, event.y_root
        if not self._pointer_inside_canvas(xr, yr):
            return
        if self._suppress_canvas_wheel(event):
            return
        self._do_canvas_wheel_linux(direction)

    def _bind_multiline_scroll_forwards_canvas(self) -> None:
        def win_wheel(e) -> str | None:
            d = getattr(e, "delta", 0) or 0
            if d:
                self._do_canvas_wheel_delta(d)
            return "break"

        def linux_up(_e) -> str | None:
            self._do_canvas_wheel_linux(-1)
            return "break"

        def linux_down(_e) -> str | None:
            self._do_canvas_wheel_linux(1)
            return "break"

        for w in (
            self.txt_pre,
            self.txt_action_observed,
            self.txt_note,
            self.txt_expected,
            self.txt_recovery,
        ):
            w.bind("<MouseWheel>", win_wheel)
            if sys.platform.startswith("linux"):
                w.bind("<Button-4>", linux_up)
                w.bind("<Button-5>", linux_down)

    def _bind_description_tab_order(self) -> None:
        """Precondition → … → Recovery → Error rate → Attachment → Contact info."""
        chain = (
            self.txt_pre,
            self.txt_action_observed,
            self.txt_note,
            self.txt_expected,
            self.txt_recovery,
            self._entry_error_rate,
            self._entry_attach,
            self._entry_contact_info,
        )

        def focus_next(target):
            def _on_tab(_event):
                target.focus_set()
                if target in (self.txt_pre, self.txt_action_observed):
                    self._move_text_cursor_to_first_number(target)
                return "break"

            return _on_tab

        def focus_prev(target):
            def _on_shift_tab(_event):
                target.focus_set()
                if target in (self.txt_pre, self.txt_action_observed):
                    self._move_text_cursor_to_first_number(target)
                return "break"

            return _on_shift_tab

        for i in range(len(chain) - 1):
            chain[i].bind("<Tab>", focus_next(chain[i + 1]))
        for i in range(1, len(chain)):
            chain[i].bind("<Shift-Tab>", focus_prev(chain[i - 1]))
            chain[i].bind("<ISO_Left_Tab>", focus_prev(chain[i - 1]))

    @staticmethod
    def _clear_text_selection(w: Text) -> None:
        try:
            w.tag_remove("sel", "1.0", END)
        except Exception:
            pass

    @staticmethod
    def _move_text_cursor_to_first_number(w: Text) -> None:
        BugReportApp._clear_text_selection(w)
        content = w.get("1.0", END)
        pos = content.find("1.")
        if pos >= 0:
            w.mark_set("insert", f"1.0+{pos + 2}c")
        else:
            w.mark_set("insert", "1.0")
        w.see("insert")

    def _build_scrollable_form(self) -> None:
        outer = Frame(self.root)
        outer.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        canvas = Canvas(outer, highlightthickness=0, yscrollincrement=24)
        self._scroll_canvas = canvas

        def _vsb_scroll(*args):
            self._unpost_all_ttk_comboboxes()
            canvas.yview(*args)

        vsb = ttk.Scrollbar(outer, orient="vertical", command=_vsb_scroll)
        form = Frame(canvas)
        self._scroll_form = form
        win = canvas.create_window((0, 0), window=form, anchor="nw")

        self._build_excel_bar(form)

        def _on_form_configure(_event=None) -> None:
            canvas.after_idle(self._refresh_scrollregion)

        def _canvas_width(event) -> None:
            canvas.itemconfigure(win, width=event.width)

        form.bind("<Configure>", _on_form_configure)
        canvas.bind("<Configure>", _canvas_width)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.root.bind_all("<MouseWheel>", self._wheel_scroll)
        if sys.platform.startswith("linux"):
            self.root.bind_all("<Button-4>", lambda e: self._wheel_scroll_linux(e, -1))
            self.root.bind_all("<Button-5>", lambda e: self._wheel_scroll_linux(e, 1))

        self.var_title = StringVar()
        self.var_full_text = StringVar()
        self.var_reporter = StringVar()
        self.var_country = StringVar(value=DEFAULT_COUNTRY)
        self.var_found_date = StringVar()
        self.var_found_time = StringVar()
        self.var_rating = StringVar(value="A")
        self.var_function = StringVar(value="")
        self.var_function_detail = StringVar(value="")
        self.var_frequency = StringVar(value="Always")
        self.var_project = StringVar(value=DEFAULT_PROJECT)
        self._env_vars = {k: StringVar() for k in sorted(ENV_STORAGE_KEYS)}
        self._vehicle_bench_option_list: list[str] = []
        self._vehicle_bench_row_attrs: dict[str, dict[str, str]] = {}
        self._vehicle_bench_resync_job: str | int | None = None
        self._function_option_list: list[str] = []
        self._function_to_details: dict[str, list[str]] = {}
        self._cb_function: ttk.Combobox | None = None
        self._cb_function_detail: ttk.Combobox | None = None
        self._env_line1_widgets: list[tuple[Label, ttk.Entry | ttk.Combobox]] = []
        self._env_line2_widgets: list[tuple[Label, ttk.Entry]] = []
        self._env_line3_widgets: list[tuple[Label, ttk.Entry]] = []
        self._env_blank_var = StringVar(value="")
        self._entry_db: ttk.Entry | None = None
        self._entry_apk: ttk.Entry | None = None
        self._env_entry_widgets: dict[str, ttk.Entry] = {}
        self.var_error_rate = StringVar()
        self.var_contact_info = StringVar()
        self.var_attach = StringVar(value=DEFAULT_ATTACHMENT)
        self.var_known_issue = BooleanVar(value=False)
        self.var_spec_check = BooleanVar(value=False)
        self.var_retest_verification = BooleanVar(value=False)
        self.var_kpm_number = StringVar(value="")

        sec_basic = ttk.LabelFrame(form, text="Basic", padding=10)
        sec_basic.pack(fill="x", pady=(0, 8))
        self._last_title_preview_values: tuple[str, ...] | None = None
        self._suppress_title_preview_nav = False
        title_row = Frame(sec_basic)
        title_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=4)
        sec_basic.columnconfigure(1, weight=1)
        Label(
            title_row,
            text=f"Title * (max {TITLE_MAX_LEN} chars)",
            anchor="w",
        ).pack(side="left", anchor="n")
        self._entry_title = ttk.Entry(
            title_row, textvariable=self.var_title, width=TITLE_ENTRY_VISIBLE_WIDTH
        )
        self._entry_title.pack(side="left", padx=(8, 0), anchor="n")
        vcmd_title = (self.root.register(lambda p: len(p) <= TITLE_MAX_LEN), "%P")
        self._entry_title.configure(validate="key", validatecommand=vcmd_title)

        def _on_title_maxlen_focusout(_event=None) -> None:
            s = self.var_title.get()
            if len(s) > TITLE_MAX_LEN:
                self.var_title.set(s[:TITLE_MAX_LEN])

        self._entry_title.bind("<FocusOut>", _on_title_maxlen_focusout)

        pv_wrap = Frame(title_row)
        pv_wrap.pack(side="left", padx=(10, 0), anchor="n")
        Label(pv_wrap, text="Preview", anchor="w").pack(side="left", padx=(0, 8), anchor="center")
        self._cb_title_preview = ttk.Combobox(
            pv_wrap,
            state="readonly",
            width=TITLE_PREVIEW_COMBO_WIDTH,
            values=(),
        )
        self._cb_title_preview.pack(side="left", anchor="center")
        self._cb_title_preview.bind("<<ComboboxSelected>>", self._on_title_preview_selected)

        self.var_title.trace_add("write", lambda *_: self._refresh_title_preview())
        ft_block = Frame(sec_basic)
        ft_block.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        Label(ft_block, text="Full Text", anchor="w").pack(side="left", padx=(0, 8))
        ttk.Entry(ft_block, textvariable=self.var_full_text, width=FORM_WIDE_WIDTH).pack(side="left")
        self._row_date_timestamps_reporter_basic(sec_basic, 2)
        self._row_rating_function_basic(sec_basic, 3)

        sec_env = ttk.LabelFrame(form, text="Environment", padding=(0, 8, 8, 8))
        sec_env.pack(fill="x", pady=(0, 8))
        self._build_environment_section(sec_env)

        sec_tracking = ttk.LabelFrame(form, text="Issue flags / KPM", padding=10)
        sec_tracking.pack(fill="x", pady=(0, 8))
        row_flags = Frame(sec_tracking)
        row_flags.pack(fill="x", anchor="w")
        Label(row_flags, text="KPM", anchor="w").pack(side="left", padx=(0, 8))
        self._entry_kpm_number = ttk.Entry(
            row_flags,
            textvariable=self.var_kpm_number,
            width=16,
        )
        self._entry_kpm_number.pack(side="left", padx=(0, 20))

        def _on_kpm_focusout(_event=None) -> None:
            s = self.var_kpm_number.get()
            if len(s) > KPM_TEXT_MAX_LEN:
                self.var_kpm_number.set(s[:KPM_TEXT_MAX_LEN])

        self._entry_kpm_number.bind("<FocusOut>", _on_kpm_focusout)

        ttk.Checkbutton(row_flags, text="Known issue", variable=self.var_known_issue).pack(
            side="left", padx=(0, 20)
        )
        ttk.Checkbutton(row_flags, text="Spec check", variable=self.var_spec_check).pack(
            side="left", padx=(0, 20)
        )
        ttk.Checkbutton(
            row_flags, text="Retest/Verification", variable=self.var_retest_verification
        ).pack(side="left")

        sec_detail = ttk.LabelFrame(form, text="Description", padding=10)
        sec_detail.pack(fill="x", pady=(0, 8))
        self.txt_pre = self._multiline_packed(
            sec_detail,
            "Precondition",
            PRECONDITION_MULTILINE_HEIGHT,
            width=DESCRIPTION_MULTILINE_WIDTH,
        )
        self.txt_action_observed = self._multiline_packed(
            sec_detail, "Action and Observation", 12, width=DESCRIPTION_MULTILINE_WIDTH
        )
        self.txt_note = self._multiline_packed(sec_detail, "Note", 1, width=DESCRIPTION_MULTILINE_WIDTH)
        self.txt_expected = self._multiline_packed(
            sec_detail, "Expected behavior", 1, width=DESCRIPTION_MULTILINE_WIDTH
        )
        self.txt_recovery = self._multiline_packed(sec_detail, "Recovery", 1, width=DESCRIPTION_MULTILINE_WIDTH)
        er_attach = Frame(sec_detail)
        er_attach.pack(fill="x", pady=(0, 8))
        Label(er_attach, text="Error rate", anchor="w").pack(side="left")
        self._entry_error_rate = ttk.Entry(
            er_attach, textvariable=self.var_error_rate, width=FORM_NARROW_WIDTH
        )
        self._entry_error_rate.pack(side="left", padx=(4, 20))
        Label(er_attach, text="Attachment", anchor="w").pack(side="left")
        self._entry_attach = ttk.Entry(
            er_attach, textvariable=self.var_attach, width=FORM_NARROW_WIDTH
        )
        self._entry_attach.pack(side="left", padx=(4, 0))
        contact_row = Frame(sec_detail)
        contact_row.pack(fill="x", pady=(0, 8))
        Label(contact_row, text="Contact info", anchor="w").pack(side="left", padx=(0, 4))
        self._entry_contact_info = ttk.Entry(
            contact_row, textvariable=self.var_contact_info, width=CONTACT_INFO_ENTRY_WIDTH
        )
        self._entry_contact_info.pack(side="left")

        for w in (self.txt_pre, self.txt_action_observed):
            w.bind("<FocusIn>", lambda _e, ww=w: self._clear_text_selection(ww), add="+")

        self._bind_multiline_scroll_forwards_canvas()
        self._bind_description_tab_order()
        self.root.bind_all("<FocusIn>", self._on_focus_in_scroll_into_view, add="+")
        canvas.after(100, self._refresh_scrollregion)

    def _row_labeled_entry(
        self,
        parent: Frame,
        row: int,
        label: str,
        var: StringVar,
        width: int = 50,
        max_length: int | None = None,
        entry_sticky: str = "ew",
    ) -> None:
        Label(parent, text=label, anchor="w").grid(row=row, column=0, sticky="nw", pady=4)
        ent = ttk.Entry(parent, textvariable=var, width=width)
        if max_length is not None:
            # Truncate on FocusOut only: trace("write") runs during IME preedit and can feel sluggish.

            def _on_maxlen_focusout(_event=None, _lim: int = max_length, _v: StringVar = var) -> None:
                s = _v.get()
                if len(s) > _lim:
                    _v.set(s[:_lim])

            ent.bind("<FocusOut>", _on_maxlen_focusout)
        ent.grid(row=row, column=1, sticky=entry_sticky, padx=(8, 0), pady=4)
        parent.columnconfigure(1, weight=1)

    def _on_project_change(self) -> None:
        self._refresh_environment_rows()

    def _on_append_checkbox_return(self, _event=None) -> str | None:
        self.append_mode.set(not self.append_mode.get())
        return "break"

    def _on_project_label_return(self, _event=None) -> str | None:
        ids = list(PROJECT_IDS)
        cur = _normalize_project(self.var_project.get())
        try:
            i = ids.index(cur)
        except ValueError:
            i = 0
        self.var_project.set(ids[(i + 1) % len(ids)])
        self._on_project_change()
        return "break"

    def _bind_enter_runs_command(self, widget, cmd) -> None:
        def run(_event=None) -> str | None:
            cmd()
            return "break"

        widget.bind("<Return>", run)
        widget.bind("<KP_Enter>", run)

    @staticmethod
    def _widget_is_descendant_of(widget, ancestor) -> bool:
        w = widget
        while w is not None:
            if w == ancestor:
                return True
            try:
                w = w.master
            except Exception:
                break
        return False

    def _do_scroll_widget_into_view(self, w) -> None:
        """Scroll the main form canvas so ``w`` (inside ``_scroll_form``) is visible."""
        canvas = self._scroll_canvas
        form = self._scroll_form
        if canvas is None or form is None:
            return
        try:
            if not w.winfo_exists():
                return
        except Exception:
            return
        if not self._widget_is_descendant_of(w, form):
            return
        try:
            w.update_idletasks()
        except Exception:
            return
        canvas.update_idletasks()
        self._refresh_scrollregion()
        canvas.update_idletasks()
        bbox = canvas.bbox("all")
        if not bbox:
            return
        y0, total_bottom = bbox[1], bbox[3]
        total_h = max(1.0, float(total_bottom - y0))
        y_acc = 0.0
        cur = w
        while cur is not None and cur != form:
            try:
                y_acc += float(cur.winfo_y())
            except Exception:
                return
            cur = cur.master
        y_top = y_acc
        try:
            y_bot = y_top + max(1, int(w.winfo_height()))
        except Exception:
            y_bot = y_top + 1
        w_top_fr = y_top / total_h
        w_bot_fr = y_bot / total_h
        try:
            vis = canvas.yview()
            top_fr = float(vis[0])
            bot_fr = float(vis[1])
        except Exception:
            return
        view_h_fr = max(0.001, bot_fr - top_fr)
        margin = min(0.02, view_h_fr * 0.1)
        if w_top_fr < top_fr + margin:
            canvas.yview_moveto(max(0.0, w_top_fr - margin))
        elif w_bot_fr > bot_fr - margin:
            canvas.yview_moveto(max(0.0, min(1.0 - view_h_fr, w_bot_fr - view_h_fr + margin)))

    def _on_focus_in_scroll_into_view(self, event) -> None:
        w = event.widget
        if self._scroll_form is None or self._scroll_canvas is None:
            return
        if not self._widget_is_descendant_of(w, self._scroll_form):
            return
        if self._focus_scroll_after_id is not None:
            try:
                self.root.after_cancel(self._focus_scroll_after_id)
            except Exception:
                pass
        self._focus_scroll_after_id = self.root.after(
            60, lambda wref=w: self._run_debounced_scroll_into_view(wref)
        )

    def _run_debounced_scroll_into_view(self, w) -> None:
        self._focus_scroll_after_id = None
        self._do_scroll_widget_into_view(w)

    def _refresh_environment_rows(self) -> None:
        if not self._env_line1_widgets or not self._env_line2_widgets or not self._env_line3_widgets:
            return
        pid = _normalize_project(self.var_project.get())
        fields = PROJECT_ENV_FIELDS[pid]
        for i in range(3):
            key, label_text = fields[i]
            lbl, w = self._env_line1_widgets[i]
            lbl.configure(text=label_text)
            w.configure(textvariable=self._env_vars[key])
        self._refresh_vehicle_bench_combobox_widget()
        detail_pairs = fields[3:]
        label_by_key = {k: t for k, t in detail_pairs}
        moved_keys = {
            # MIB3: MU + Conbox+GW on one row. HCP3/ICC/MEB: HCP3 or ICC + Conmod (+ GW/HCP5/ICAS2) on one row.
            "mu_hw",
            "mu_sw",
            "hcp3_hw",
            "hcp3_sw",
            "icc_hw",
            "icc_sw",
            "conbox_hw",
            "conbox_sw",
            "conmod_hw",
            "conmod_sw",
            "hcp5_hw",
            "hcp5_sw",
            "gw_hw",
            "gw_sw",
            "icas2_hw",
            "icas2_sw",
        }
        line3_keys = [k for k, _ in detail_pairs if k in moved_keys]
        line2_keys = [k for k, _ in detail_pairs if k not in moved_keys]

        def bind_env_line(widgets: list[tuple[Label, ttk.Entry]], keys: list[str]) -> None:
            # Hide unused slots completely so trailing blank cells are not visible.
            for lbl, _ent in widgets:
                cell = lbl.master
                if cell.winfo_manager():
                    cell.pack_forget()
            for i, key in enumerate(keys):
                if i >= len(widgets):
                    break
                lbl, ent = widgets[i]
                cell = lbl.master
                lbl.configure(text=label_by_key.get(key, ""))
                ent.configure(state="normal", textvariable=self._env_vars[key])
                self._env_entry_widgets[key] = ent
                if key in {"mu_sw", "hcp3_sw", "icc_sw"}:
                    ent.configure(width=ENV_LINE2_SW_ENTRY_WIDTH)
                else:
                    ent.configure(width=ENV_LINE2_ENTRY_WIDTH)
                right_pad = ENV_LINE2_GROUP_GAP if i < len(keys) - 1 else 0
                cell.pack(side="left", padx=(0, right_pad), pady=2)

        self._env_entry_widgets = {}
        bind_env_line(self._env_line2_widgets, line2_keys)
        bind_env_line(self._env_line3_widgets, line3_keys)
        self._rebind_environment_tab_order(line2_keys, line3_keys)

    def _rebind_environment_tab_order(self, line2_keys: list[str], line3_keys: list[str]) -> None:
        if self._entry_db is None or self._entry_apk is None:
            return
        # Reset to default tab behavior first.
        self._entry_db.bind("<Tab>", lambda _e: None)
        self._entry_apk.bind("<Tab>", lambda _e: None)
        for ent in self._env_entry_widgets.values():
            ent.bind("<Tab>", lambda _e: None)

        # Apply the same DB-driven tab flow for every project:
        # DB -> APK -> visible environment fields (top row first, then bottom row).
        chain = [*line2_keys, *line3_keys]
        visible = [k for k in chain if k in self._env_entry_widgets]

        def _focus_target(widget):
            def _on_tab(_event):
                widget.focus_set()
                return "break"

            return _on_tab

        if not visible:

            def _apk_tab_to_pre(_event):
                self.txt_pre.focus_set()
                self._move_text_cursor_to_first_number(self.txt_pre)
                return "break"

            self._entry_db.bind("<Tab>", _focus_target(self._entry_apk))
            self._entry_apk.bind("<Tab>", _apk_tab_to_pre)
            return

        self._entry_db.bind("<Tab>", _focus_target(self._entry_apk))
        self._entry_apk.bind("<Tab>", _focus_target(self._env_entry_widgets[visible[0]]))
        for i in range(len(visible) - 1):
            cur = self._env_entry_widgets[visible[i]]
            nxt = self._env_entry_widgets[visible[i + 1]]
            cur.bind("<Tab>", _focus_target(nxt))
        # Last environment field -> Precondition, with cursor on the first numbered line.
        self._env_entry_widgets[visible[-1]].bind(
            "<Tab>",
            lambda _e: (
                self.txt_pre.focus_set(),
                self._move_text_cursor_to_first_number(self.txt_pre),
                "break",
            )[-1],
        )

    def _env_labeled_entry_cell(
        self,
        line: Frame,
        *,
        entry_width: int,
        group_right_pad: int | None = None,
    ) -> tuple[Label, ttk.Entry]:
        right_pad = ENV_FIELD_GROUP_GAP if group_right_pad is None else group_right_pad
        cell = Frame(line)
        cell.pack(side="left", padx=(0, right_pad), pady=2)
        lbl = Label(cell, text="", anchor="w")
        lbl.pack(side="left", padx=(0, ENV_LABEL_TO_ENTRY_GAP), pady=0)
        ent = ttk.Entry(cell, width=entry_width)
        ent.pack(side="left")
        return lbl, ent

    def _env_labeled_vehicle_bench_combobox_cell(
        self,
        line: Frame,
        *,
        width: int,
        group_right_pad: int | None = None,
    ) -> tuple[Label, ttk.Combobox]:
        right_pad = ENV_FIELD_GROUP_GAP if group_right_pad is None else group_right_pad
        cell = Frame(line)
        cell.pack(side="left", padx=(0, right_pad), pady=2)
        lbl = Label(cell, text="", anchor="w")
        lbl.pack(side="left", padx=(0, ENV_LABEL_TO_ENTRY_GAP), pady=0)
        cb = ttk.Combobox(
            cell,
            textvariable=self._env_vars["vehicle_bench"],
            values=[],
            width=width,
            state="normal",
        )
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda _e: self._on_vehicle_bench_selected())
        cb.bind("<FocusOut>", lambda _e: self._maybe_sync_vin_after_bench_edit())
        cb.bind("<FocusIn>", lambda _e: self._schedule_vehicle_bench_resync_from_excel())
        cb.bind("<Button-1>", lambda _e: self._schedule_vehicle_bench_resync_from_excel())
        return lbl, cb

    def _vehicle_bench_dropdown_values(self) -> list[str]:
        opts = list(self._vehicle_bench_option_list)
        cur = (self._env_vars["vehicle_bench"].get() or "").strip()
        if cur and cur not in opts:
            return [cur, *opts]
        return opts

    def _refresh_vehicle_bench_combobox_widget(self) -> None:
        if len(self._env_line1_widgets) < 2:
            return
        _, w = self._env_line1_widgets[1]
        if not isinstance(w, ttk.Combobox):
            return
        vals = self._vehicle_bench_dropdown_values()
        w["values"] = vals
        # Keep editable: Project can be chosen first; user may type a bench not in general_info.xlsx.
        w.configure(state="normal")

    def _sync_vehicle_row_from_general_info(self) -> None:
        """Apply Project, Cluster, VIN from the same general_info row as the selected Vehicle/Bench."""
        vb = (self._env_vars["vehicle_bench"].get() or "").strip()
        if not vb:
            return
        attrs = self._vehicle_bench_row_attrs.get(vb)
        if not attrs:
            return
        applied = False
        for key in ("cluster", "vin"):
            val = (attrs.get(key) or "").strip()
            if not val:
                continue
            self._env_vars[key].set(val)
            self._reports[self._current_idx][key] = val
            applied = True
        pv = (attrs.get("project") or "").strip()
        if pv:
            norm = _normalize_project(pv)
            self.var_project.set(norm)
            self._session_common["project"] = norm
            for rep in self._reports:
                rep["project"] = norm
            applied = True
        if applied:
            self._refresh_environment_rows()

    def _on_vehicle_bench_selected(self) -> None:
        self._sync_vehicle_row_from_general_info()

    def _maybe_sync_vin_after_bench_edit(self) -> None:
        """When Vehicle/Bench is editable, apply row mapping after the field loses focus."""
        self._sync_vehicle_row_from_general_info()

    def _schedule_vehicle_bench_resync_from_excel(self) -> None:
        """Re-read general_info shortly before the user opens the list (external Excel edits)."""
        if self._vehicle_bench_resync_job is not None:
            try:
                self.root.after_cancel(self._vehicle_bench_resync_job)
            except Exception:
                pass
        self._vehicle_bench_resync_job = self.root.after(
            80, self._run_vehicle_bench_resync_from_excel
        )

    def _run_vehicle_bench_resync_from_excel(self) -> None:
        self._vehicle_bench_resync_job = None
        self._reload_vehicle_bench_options_from_general_info()

    def _reload_vehicle_bench_options_from_general_info(self) -> None:
        p = resolve_general_info_excel_path(self.default_info_path.get())
        self._vehicle_bench_option_list, self._vehicle_bench_row_attrs = _read_vehicle_bench_sync_from_excel(p)
        self._refresh_vehicle_bench_combobox_widget()
        self._sync_vehicle_row_from_general_info()

    def _build_environment_section(self, parent: Frame) -> None:
        parent.columnconfigure(0, weight=1)

        # One row: Project + radios stay together (wide col0 from line1/line2 would hide column 1).
        project_row = Frame(parent)
        project_row.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        lbl_project = Label(project_row, text="Project", anchor="w", takefocus=1)
        lbl_project.pack(side="left", padx=(0, 8))
        lbl_project.bind("<Return>", self._on_project_label_return)
        lbl_project.bind("<KP_Enter>", self._on_project_label_return)
        for pid in PROJECT_IDS:
            rb = Radiobutton(
                project_row,
                text=pid,
                variable=self.var_project,
                value=pid,
                command=self._on_project_change,
                takefocus=1,
            )
            rb.pack(side="left", padx=(0, 6))

            def on_project_radio_return(_e, p: str = pid) -> str | None:
                self.var_project.set(p)
                self._on_project_change()
                return "break"

            rb.bind("<Return>", on_project_radio_return)
            rb.bind("<KP_Enter>", on_project_radio_return)

        # Full-width rows, left-aligned with Project (same insets as project_row).
        line1 = Frame(parent)
        line1.grid(row=1, column=0, sticky="w", pady=(4, 6))
        for i in range(3):
            if i == 1:
                self._env_line1_widgets.append(
                    self._env_labeled_vehicle_bench_combobox_cell(
                        line1,
                        width=ENV_ENTRY_WIDTH,
                        group_right_pad=ENV_LINE1_GROUP_GAP,
                    )
                )
            else:
                self._env_line1_widgets.append(
                    self._env_labeled_entry_cell(
                        line1,
                        entry_width=ENV_VIN_ENTRY_WIDTH if i == 2 else ENV_ENTRY_WIDTH,
                        group_right_pad=ENV_LINE1_GROUP_GAP if i < 2 else 0,
                    )
                )

        line2 = Frame(parent)
        line2.grid(row=2, column=0, sticky="w", pady=(0, 4))
        for i in range(6):
            self._env_line2_widgets.append(
                self._env_labeled_entry_cell(
                    line2,
                    entry_width=ENV_LINE2_ENTRY_WIDTH,
                    group_right_pad=ENV_LINE2_GROUP_GAP if i < 5 else 0,
                )
            )

        db_cell = Frame(line2)
        db_cell.pack(side="left", padx=(0, 0), pady=2)
        Label(db_cell, text="DB", anchor="w").pack(side="left", padx=(0, ENV_LABEL_TO_ENTRY_GAP), pady=0)
        self._entry_db = ttk.Entry(db_cell, textvariable=self._env_vars["db"], width=ENV_ENTRY_WIDTH * 2)
        self._entry_db.pack(side="left")

        apk_cell = Frame(line2)
        apk_cell.pack(side="left", padx=(12, 0), pady=2)
        Label(apk_cell, text="APK", anchor="w").pack(side="left", padx=(0, ENV_LABEL_TO_ENTRY_GAP), pady=0)
        self._entry_apk = ttk.Entry(apk_cell, textvariable=self._env_vars["apk"], width=ENV_ENTRY_WIDTH * 2)
        self._entry_apk.pack(side="left")

        line3 = Frame(parent)
        line3.grid(row=3, column=0, sticky="w", pady=(0, 4))
        for i in range(6):
            self._env_line3_widgets.append(
                self._env_labeled_entry_cell(
                    line3,
                    entry_width=ENV_LINE2_ENTRY_WIDTH,
                    group_right_pad=ENV_LINE2_GROUP_GAP if i < 5 else 0,
                )
            )

        # Custom tab order: VIN -> DB -> APK first.
        if len(self._env_line1_widgets) >= 3 and self._entry_db is not None:
            vin_widget = self._env_line1_widgets[2][1]
            if isinstance(vin_widget, ttk.Entry):
                def _on_vin_tab(_event):
                    if self._entry_db is not None:
                        self._entry_db.focus_set()
                    return "break"

                vin_widget.bind("<Tab>", _on_vin_tab)

        self._refresh_environment_rows()

    def _row_combobox(
        self,
        parent: Frame,
        row: int,
        label: str,
        var: StringVar,
        values: list[str],
        width: int,
    ) -> None:
        Label(parent, text=label, anchor="w").grid(row=row, column=0, sticky="nw", pady=4)
        cb = ttk.Combobox(
            parent,
            textvariable=var,
            values=values,
            state="readonly",
            width=width,
        )
        cb.grid(row=row, column=1, sticky="w", padx=(8, 0), pady=4)

    def _row_date_timestamps_reporter_basic(self, parent: Frame, row: int) -> None:
        inner = Frame(parent)
        inner.grid(row=row, column=0, columnspan=2, sticky="ew", pady=4)
        Label(inner, text="Date").pack(side="left")
        ttk.Entry(inner, textvariable=self.var_found_date, width=14).pack(side="left", padx=(4, 12))
        Label(inner, text="Timestamps (HH:MM)").pack(side="left")
        ttk.Entry(inner, textvariable=self.var_found_time, width=10).pack(side="left", padx=(4, 12))
        Label(inner, text="Reporter *").pack(side="left")
        ttk.Entry(inner, textvariable=self.var_reporter, width=REPORTER_ENTRY_WIDTH).pack(
            side="left", padx=(4, 0)
        )
        Label(inner, text="Country").pack(side="left", padx=(12, 0))
        ttk.Combobox(
            inner,
            textvariable=self.var_country,
            values=COUNTRY_COMBO_VALUES,
            width=8,
            state="normal",
        ).pack(side="left", padx=(4, 0))
        parent.columnconfigure(1, weight=1)

    def _row_rating_function_basic(self, parent: Frame, row: int) -> None:
        inner = Frame(parent)
        inner.grid(row=row, column=0, columnspan=2, sticky="w", pady=4)
        Label(inner, text="Rating").pack(side="left")
        ttk.Combobox(
            inner,
            textvariable=self.var_rating,
            values=["A1", "A", "B", "C"],
            state="readonly",
            width=6,
        ).pack(side="left", padx=(4, 12))
        Label(inner, text="Frequency").pack(side="left")
        ttk.Combobox(
            inner,
            textvariable=self.var_frequency,
            values=["Always", "Sometimes", "Once"],
            state="readonly",
            width=12,
        ).pack(side="left", padx=(4, 12))
        Label(inner, text="Function").pack(side="left")
        self._cb_function = ttk.Combobox(
            inner,
            textvariable=self.var_function,
            values=[],
            width=14,
            state="normal",
        )
        self._cb_function.pack(side="left", padx=(4, 12))
        self._cb_function.bind("<<ComboboxSelected>>", lambda _e: self._on_function_selected())
        self._cb_function.bind("<FocusOut>", lambda _e: self._maybe_refresh_function_detail_after_function_edit())
        Label(inner, text="Function detail").pack(side="left")
        self._cb_function_detail = ttk.Combobox(
            inner,
            textvariable=self.var_function_detail,
            values=[],
            width=22,
            state="normal",
        )
        self._cb_function_detail.pack(side="left", padx=(4, 0))
        parent.columnconfigure(1, weight=1)

    def _packed_combobox_row(
        self,
        parent: Frame,
        label: str,
        var: StringVar,
        values: list[str],
        width: int,
    ) -> None:
        rowf = Frame(parent)
        rowf.pack(fill="x", pady=(0, 8))
        Label(rowf, text=label, anchor="w").pack(anchor="w")
        ttk.Combobox(
            rowf,
            textvariable=var,
            values=values,
            state="readonly",
            width=width,
        ).pack(anchor="w", pady=(4, 0))

    def _multiline_packed(
        self,
        parent: Frame,
        title: str,
        height: int,
        width: int | None = None,
    ) -> Text:
        block = Frame(parent)
        block.pack(fill="x", pady=(0, 8))
        Label(block, text=title, anchor="w").pack(anchor="w")
        t = Text(block, height=height, wrap="word", font=("Segoe UI", 10))
        if width is not None:
            t.configure(width=width)
        t.pack(anchor="w", pady=(4, 0))
        return t

    def _packed_labeled_entry(
        self,
        parent: Frame,
        label: str,
        var: StringVar,
        width: int = 50,
    ) -> None:
        row = Frame(parent)
        row.pack(fill="x", pady=(0, 8))
        Label(row, text=label, anchor="w").pack(anchor="w")
        ttk.Entry(row, textvariable=var, width=width).pack(anchor="w", pady=(4, 0))

    def _build_actions(self) -> None:
        bottom = Frame(self.root, padx=8, pady=8)
        bottom.pack(fill="x")
        btn_save = ttk.Button(bottom, text="Save all reports to Excel", command=self._save_excel)
        btn_save.pack(side="left")
        self._bind_enter_runs_command(btn_save, self._save_excel)
        btn_reset = ttk.Button(bottom, text="Reset current report", command=self._reset_form)
        btn_reset.pack(side="left", padx=(8, 0))
        self._bind_enter_runs_command(btn_reset, self._reset_form)
        self._btn_delete_report = ttk.Button(
            bottom,
            text="Delete report",
            command=self._delete_current_report,
        )
        self._btn_delete_report.pack(side="left", padx=(8, 0))
        self._bind_enter_runs_command(self._btn_delete_report, self._delete_current_report)
        ttk.Button(bottom, text="Exit", command=self._quit_app).pack(side="right")

    def _update_nav_label(self) -> None:
        n = len(self._reports)
        self._nav_label_var.set(f"Report {self._current_idx + 1} / {n}")
        self._sync_nav_prev_button()
        self._refresh_title_preview()

    @staticmethod
    def _preview_title_short(raw: str, max_len: int = TITLE_PREVIEW_LINE_CHARS) -> str:
        t = (raw or "").replace("\n", " ").strip()
        if len(t) <= max_len:
            return t
        return t[: max_len - 1] + "\u2026"

    def _title_preview_values_tuple(self) -> tuple[str, ...]:
        cur = self._current_idx
        out: list[str] = []
        for i, rep in enumerate(self._reports):
            if i == cur:
                t = self.var_title.get()
            else:
                t = rep.get("title", "")
            short = self._preview_title_short(t)
            prefix = f"{i + 1}. "
            if short:
                out.append(prefix + short)
            else:
                out.append(prefix + "(no title)")
        return tuple(out)

    def _refresh_title_preview(self) -> None:
        if not hasattr(self, "_cb_title_preview"):
            return
        vals = self._title_preview_values_tuple()
        if vals == getattr(self, "_last_title_preview_values", None):
            # Still move selection if index changed with same labels (e.g. after nav).
            self._suppress_title_preview_nav = True
            try:
                if 0 <= self._current_idx < len(vals):
                    self._cb_title_preview.current(self._current_idx)
            finally:
                self._suppress_title_preview_nav = False
            return
        self._last_title_preview_values = vals
        self._suppress_title_preview_nav = True
        try:
            self._cb_title_preview["values"] = vals
            if vals and 0 <= self._current_idx < len(vals):
                self._cb_title_preview.current(self._current_idx)
            elif vals:
                self._cb_title_preview.current(0)
            else:
                self._cb_title_preview.set("")
        finally:
            self._suppress_title_preview_nav = False

    def _on_title_preview_selected(self, _event=None) -> None:
        if getattr(self, "_suppress_title_preview_nav", False):
            return
        cb = self._cb_title_preview
        idx = cb.current()
        if idx < 0 or idx >= len(self._reports):
            return
        if idx == self._current_idx:
            return
        self._flush_current()
        self._current_idx = idx
        self._apply_to_form(self._reports[idx])
        self._update_nav_label()
        self.root.after_idle(self._refresh_scrollregion)

    def _on_nav_prev_clicked(self, _event=None) -> str | None:
        if self._current_idx <= 0:
            return self._on_first_page_prev_notice(_event)
        self._nav_prev()
        return "break" if _event is not None else None

    def _on_first_page_prev_notice(self, _event=None) -> str | None:
        messagebox.showinfo(
            "No previous report",
            "You are on the first report. There is no previous page.",
        )
        return "break"

    def _sync_nav_prev_button(self) -> None:
        if not hasattr(self, "_btn_nav_prev"):
            return
        self._btn_nav_prev.configure(state="normal")
        self._bind_enter_runs_command(self._btn_nav_prev, self._on_nav_prev_clicked)
        try:
            if self._current_idx <= 0:
                self._btn_nav_prev.configure(style="NavPrevMuted.TButton")
            else:
                self._btn_nav_prev.configure(style="TButton")
        except Exception:
            pass

    def _delete_current_report(self) -> None:
        if len(self._reports) <= 1:
            messagebox.showinfo(
                "Cannot delete report",
                "This is the only report. It cannot be deleted.",
            )
            return
        if not messagebox.askyesno(
            "Delete report",
            f"Delete report {self._current_idx + 1} of {len(self._reports)}? This cannot be undone.",
        ):
            return
        self._flush_current()
        del self._reports[self._current_idx]
        if self._current_idx >= len(self._reports):
            self._current_idx = len(self._reports) - 1
        self._apply_to_form(self._reports[self._current_idx])
        self._update_nav_label()
        self.root.after_idle(self._refresh_scrollregion)

    def _snapshot_from_form(self) -> dict[str, str]:
        self._sync_common_from_form()
        act, obs = _split_action_observation_combined(
            self._get_multiline(self.txt_action_observed)
        )
        return {
            "title": self.var_title.get().strip()[:TITLE_MAX_LEN],
            "full_text": self.var_full_text.get().strip(),
            "reporter": self._session_common["reporter"],
            "country": self._session_common.get("country", "") or DEFAULT_COUNTRY,
            "found_date": self.var_found_date.get().strip(),
            "found_time": self.var_found_time.get().strip(),
            "rating": self.var_rating.get().strip(),
            "function": self.var_function.get().strip(),
            "function_detail": self.var_function_detail.get().strip(),
            "project": self._session_common["project"],
            **{k: self._env_vars[k].get().strip() for k in ENV_STORAGE_KEYS},
            "precondition": self._get_multiline(self.txt_pre),
            "frequency": self.var_frequency.get().strip(),
            "action": act,
            "observed": obs,
            "expected": self._get_multiline(self.txt_expected),
            "recovery": self._get_multiline(self.txt_recovery),
            "error_rate": self.var_error_rate.get().strip(),
            "contact_info": self.var_contact_info.get().strip(),
            "attach": self.var_attach.get().strip(),
            "note": self._get_multiline(self.txt_note),
            "known_issue": "1" if self.var_known_issue.get() else "",
            "spec_check": "1" if self.var_spec_check.get() else "",
            "retest_verification": "1" if self.var_retest_verification.get() else "",
            "kpm_number": (self.var_kpm_number.get() or "").strip()[:KPM_TEXT_MAX_LEN],
        }

    def _apply_to_form(self, d: dict[str, str]) -> None:
        self._apply_common_to_form()
        for k in ENV_STORAGE_KEYS:
            self._env_vars[k].set(d.get(k, ""))
        self._refresh_environment_rows()

        fd = (d.get("found_date") or "").strip()
        ft = (d.get("found_time") or "").strip()
        if not fd:
            fd = datetime.now().strftime("%Y-%m-%d")
        if not ft:
            ft = datetime.now().strftime("%H:%M")
        self.var_found_date.set(fd)
        self.var_found_time.set(ft)

        self.var_title.set((d.get("title", "") or "")[:TITLE_MAX_LEN])
        self.var_full_text.set(d.get("full_text", ""))

        def _set_text(w: Text, s: str) -> None:
            w.delete("1.0", END)
            w.insert("1.0", s)

        rating = d.get("rating", "A") or "A"
        if rating not in ("A1", "A", "B", "C"):
            rating = "A"
        self.var_rating.set(rating)

        fn = (d.get("function") or d.get("category", "") or "").strip()
        self.var_function.set(fn)
        self.var_function_detail.set(d.get("function_detail", "") or "")

        freq = d.get("frequency", "Always") or "Always"
        if freq not in ("Always", "Sometimes", "Once"):
            freq = "Always"
        self.var_frequency.set(freq)
        self.var_error_rate.set(d.get("error_rate", ""))
        self.var_contact_info.set(d.get("contact_info", ""))
        self.var_attach.set(d.get("attach", ""))

        self.var_known_issue.set(_flag_truthy(d.get("known_issue")))
        self.var_spec_check.set(_flag_truthy(d.get("spec_check")))
        self.var_retest_verification.set(_flag_truthy(d.get("retest_verification")))
        raw_kpm = (d.get("kpm_number") or "").strip()
        self.var_kpm_number.set(raw_kpm[:KPM_TEXT_MAX_LEN])

        _set_text(self.txt_pre, _precondition_for_form_display(d.get("precondition")))
        _set_text(
            self.txt_action_observed,
            _join_action_observation_for_display(
                d.get("action", ""), d.get("observed", "")
            ),
        )
        _set_text(self.txt_expected, d.get("expected", ""))
        _set_text(self.txt_recovery, d.get("recovery", ""))
        note = d.get("note", "") or d.get("remarks", "")
        _set_text(self.txt_note, note)
        self._refresh_vehicle_bench_combobox_widget()
        self._refresh_function_combobox_widget()
        self._coerce_function_detail_for_selected_function()

    def _flush_current(self) -> None:
        self._reports[self._current_idx] = self._snapshot_from_form()

    def _nav_prev(self) -> None:
        self._flush_current()
        if self._current_idx > 0:
            self._current_idx -= 1
            self._apply_to_form(self._reports[self._current_idx])
        self._update_nav_label()
        self.root.after_idle(self._refresh_scrollregion)

    def _nav_next(self) -> None:
        self._flush_current()
        if self._current_idx >= len(self._reports) - 1:
            prev = self._reports[self._current_idx]
            new_rep = self._new_report_dict()
            # Carry over from the report you leave when appending a new slot only (existing slots keep their own data).
            p_prev = (prev.get("precondition") or "").strip()
            new_rep["precondition"] = (
                p_prev if p_prev else DEFAULT_PRECONDITION
            )
            new_rep["error_rate"] = prev.get("error_rate", "")
            new_rep["contact_info"] = prev.get("contact_info", "")
            new_rep["attach"] = (prev.get("attach") or "").strip() or DEFAULT_ATTACHMENT
            for k in ENV_STORAGE_KEYS:
                new_rep[k] = prev.get(k, "")
            self._reports.append(new_rep)
        self._current_idx += 1
        self._apply_to_form(self._reports[self._current_idx])
        self._update_nav_label()
        self.root.after_idle(self._refresh_scrollregion)

    def _reload_general_info_from_excel(self) -> None:
        """Re-read general_info.xlsx (Vehicle/Bench list, VIN map, Field/Value defaults)."""
        self._flush_current()
        self._merge_default_info_from_file(show_read_error=True)
        self._apply_to_form(self._reports[self._current_idx])
        self._sync_vehicle_row_from_general_info()
        self.root.after_idle(self._refresh_scrollregion)

    def _pick_excel(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
            initialfile=Path(self.excel_path.get()).name,
            initialdir=str(Path(self.excel_path.get()).parent),
        )
        if path:
            self.excel_path.set(path)

    def _pick_default_info_path(self) -> None:
        raw = (self.default_info_path.get() or "").strip()
        initialdir = str(Path(raw).parent) if raw else str(resource_path())
        path = filedialog.askopenfilename(
            title="Default info file",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
            initialdir=initialdir,
        )
        if path:
            self.default_info_path.set(path)
            self._flush_current()
            self._merge_default_info_from_file(show_read_error=True)
            self._apply_to_form(self._reports[self._current_idx])
            self._sync_vehicle_row_from_general_info()
            self.root.after_idle(self._refresh_scrollregion)

    def _merge_default_info_from_file(self, show_read_error: bool = False) -> None:
        ensure_bundled_general_info_excel()
        path = resolve_general_info_excel_path(self.default_info_path.get())
        self.default_info_path.set(str(path))
        data = _read_default_info_from_excel(path)
        for k in SESSION_KEYS:
            if k not in data:
                continue
            val = data[k]
            if k == "project":
                val = _normalize_project(val)
            self._session_common[k] = val
            for rep in self._reports:
                rep[k] = val
        for k in ENV_STORAGE_KEYS:
            if k not in data:
                continue
            val = data[k]
            for rep in self._reports:
                rep[k] = val
        for k in PER_REPORT_DEFAULT_INFO_KEYS:
            if k not in data:
                continue
            val = data[k]
            target = "note" if k == "remarks" else k
            for rep in self._reports:
                rep[target] = val
        if "vehicle" in data and "vehicle_bench" not in data:
            vb = data["vehicle"]
            for rep in self._reports:
                rep["vehicle_bench"] = vb
        self._reload_vehicle_bench_options_from_general_info()
        self._reload_function_options_from_general_info()
        if show_read_error and _last_general_info_read_error:
            messagebox.showwarning(
                "general_info.xlsx",
                "엑셀 파일을 읽지 못했습니다. 다른 프로그램에서 파일을 닫았는지 확인하세요.\n\n"
                + _last_general_info_read_error,
            )

    def _reload_function_options_from_general_info(self) -> None:
        p = resolve_general_info_excel_path(self.default_info_path.get())
        self._function_option_list, self._function_to_details = _read_function_sync_from_excel(p)

    def _function_dropdown_values(self) -> list[str]:
        opts = list(self._function_option_list)
        cur = (self.var_function.get() or "").strip()
        if cur and cur not in opts:
            return [cur, *opts]
        return opts

    def _refresh_function_combobox_widget(self) -> None:
        if self._cb_function is None:
            return
        self._cb_function["values"] = self._function_dropdown_values()
        # Keep normal so users can type Function by hand, not only pick from the list.
        self._cb_function.configure(state="normal")

    def _refresh_function_detail_combobox_widget(self) -> None:
        if self._cb_function_detail is None:
            return
        fn = (self.var_function.get() or "").strip()
        opts = list(self._function_to_details.get(fn, []))
        cur = (self.var_function_detail.get() or "").strip()
        vals = [cur, *opts] if cur and cur not in opts else opts
        self._cb_function_detail["values"] = vals
        # Keep normal so users can type Function detail by hand, not only pick from the list.
        self._cb_function_detail.configure(state="normal")

    def _coerce_function_detail_for_selected_function(self) -> None:
        fn = (self.var_function.get() or "").strip()
        opts = list(self._function_to_details.get(fn, []))
        if opts:
            cur = (self.var_function_detail.get() or "").strip()
            # Only default to the first Excel option when detail is still empty; do not
            # overwrite manual text that is not in the predefined list.
            if not cur:
                self.var_function_detail.set(opts[0])
        self._refresh_function_detail_combobox_widget()

    def _on_function_selected(self) -> None:
        self._coerce_function_detail_for_selected_function()

    def _maybe_refresh_function_detail_after_function_edit(self) -> None:
        self._coerce_function_detail_for_selected_function()

    def _get_multiline(self, w: Text) -> str:
        return w.get("1.0", END).rstrip("\n")

    def _validate_report(self, d: dict[str, str], index_1based: int) -> str | None:
        title = (d.get("title") or "").strip()
        reporter = (d.get("reporter") or "").strip()
        if not title:
            return f"Report {index_1based}: please enter a title."
        if not reporter:
            return f"Report {index_1based}: please enter a reporter name."
        if len(title) > TITLE_MAX_LEN:
            return f"Report {index_1based}: title must be at most {TITLE_MAX_LEN} characters."
        return None

    def _dict_to_excel_row(self, d: dict[str, str], report_no_1based: int) -> dict[str, str]:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        m = self._merge_session(d)
        return {
            "Saved_at": now,
            "Report No.": report_no_1based,
            "KPM": _format_kpm_excel_cell(m),
            "Date": m["found_date"],
            "Timestamps": m["found_time"],
            "Reporter": m["reporter"],
            "Country": (m.get("country") or "").strip() or DEFAULT_COUNTRY,
            "Vehicle": (m.get("vehicle_bench") or m.get("vehicle") or "").strip(),
            "Function": _excel_function_function_detail(m),
            "Title": m["title"],
            "Description": _format_description_excel(m),
            "Rating": m["rating"],
            "Frequency": (m.get("frequency") or "").strip() or "Always",
        }

    def _save_excel(self) -> None:
        self._flush_current()

        for i, rep in enumerate(self._reports, start=1):
            err = self._validate_report(self._merge_session(rep), i)
            if err:
                messagebox.showwarning("Validation", err)
                self._current_idx = i - 1
                self._apply_to_form(self._reports[self._current_idx])
                self._update_nav_label()
                return

        path = self._resolve_save_path()
        self.excel_path.set(str(path))

        rows = [
            self._dict_to_excel_row(r, i) for i, r in enumerate(self._reports, start=1)
        ]
        new_df = pd.DataFrame(rows, columns=COLUMNS)

        try:
            if self.append_mode.get() and path.exists():
                old = pd.read_excel(path, engine="openpyxl")
                for c in COLUMNS:
                    if c not in old.columns:
                        old[c] = ""
                old = old.reindex(columns=COLUMNS, fill_value="")
                out = pd.concat([old, new_df], ignore_index=True)
            else:
                out = new_df

            path.parent.mkdir(parents=True, exist_ok=True)
            out.to_excel(path, index=False, engine="openpyxl")
            try:
                _apply_issue_report_sheet_column_widths(path)
            except Exception:
                pass
            try:
                _add_kpm_issue_charts(path, out)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Save failed", str(e))
            return

        messagebox.showinfo(
            "Saved",
            f"Saved {len(self._reports)} report(s).\n{path.resolve()}",
        )
        # After OK: open file; small delay avoids Windows failing to launch right after a modal closes.
        self.root.after(100, lambda p=path: self._open_saved_excel(p))

    def _open_saved_excel(self, path: Path) -> None:
        p = path.resolve()
        if not p.is_file():
            return

        def _win_open_fallback() -> None:
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
            subprocess.Popen(
                ["cmd", "/c", "start", "", str(p)],
                cwd=str(p.parent),
                shell=False,
                creationflags=creationflags,
            )

        if sys.platform == "win32":
            try:
                os.startfile(str(p))
            except OSError:
                try:
                    _win_open_fallback()
                except OSError:
                    pass
            return

        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", str(p)], start_new_session=True)
            else:
                subprocess.Popen(["xdg-open", str(p)], start_new_session=True)
        except OSError:
            pass

    def _reset_form(self) -> None:
        self._sync_common_from_form()
        body = self._blank_per_report_dict()
        body.update(self._session_common)
        body.update(_empty_env_values())
        now = datetime.now()
        body["found_date"] = now.strftime("%Y-%m-%d")
        body["found_time"] = now.strftime("%H:%M")
        self._reports[self._current_idx] = body
        self._apply_to_form(body)
        self.root.after_idle(self._refresh_scrollregion)
        messagebox.showinfo(
            "Reset current report",
            "The current report has been reset.",
        )

    def _quit_app(self) -> None:
        if self._focus_scroll_after_id is not None:
            try:
                self.root.after_cancel(self._focus_scroll_after_id)
            except Exception:
                pass
            self._focus_scroll_after_id = None
        try:
            self.root.unbind_all("<MouseWheel>")
            self.root.unbind_all("<Button-4>")
            self.root.unbind_all("<Button-5>")
        except Exception:
            pass
        self.root.quit()
        self.root.destroy()

    def run(self) -> None:
        self.root.protocol("WM_DELETE_WINDOW", self._quit_app)
        self.root.mainloop()


def main() -> None:
    # Frozen exe: cwd is often wrong when launched from shortcuts; data lives next to the exe.
    if getattr(sys, "frozen", False):
        try:
            os.chdir(resource_path())
        except OSError:
            pass
    if sys.platform == "win32":
        try:
            from ctypes import windll

            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass
    app = BugReportApp()
    app.run()


if __name__ == "__main__":
    main()
